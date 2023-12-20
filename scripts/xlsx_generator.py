"""
Generate a xlsx file containing results of given methods from csv files.

Follow the main function to manage the files and parameters

Gap compute the difference on the mean of two methods and give the p-value
with a ttest.
"""
# pylint: disable=C0115,C0103
import re
import statistics
from glob import glob
from typing import Any, Iterable
from datetime import datetime

import pandas as pd  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Alignment, Font  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from openpyxl.cell.rich_text import TextBlock, CellRichText  # type: ignore
from openpyxl.cell.text import InlineFont  # type: ignore
from scipy import stats  # type: ignore
from tqdm import tqdm  # type: ignore

# import warnings


# warnings.filterwarnings("error")

# Only the P_VALUE_1 will be take in account to count the number of time a method is better.
# The two other will be colored with colors COLOR_GAP{1,2}_{1,2,3}.
P_VALUE_1 = 0.001
P_VALUE_2 = 0.05
P_VALUE_3 = 0.1
COLOR_GAP1_1 = "488f31"
COLOR_GAP1_2 = "76a263"
COLOR_GAP1_3 = "9fb494"
COLOR_GAP2_1 = "de425b"
COLOR_GAP2_2 = "dd757d"
COLOR_GAP2_3 = "d69fa1"

# When a best score is equal to the best known score : red
COLOR_BEST = "ff0000"
# When a score is the best one on the line but not best know score : orange
COLOR_BEST_FOUND = "FF7F00"
# When a best score is proven optimal : blue
COLOR_OPTIMAL = "0000ff"
# When a best score is better than the best known score : green
COLOR_NEW_BEST = "00ff00"


def main():
    """
    Choose the methods and instances and create the xlsx file
    """
    # Add method name and repertory of data of each method
    methods: list[tuple(str, str)] = [
        #
        # WVCP bound vs no bound
        # ("MCTS+R no bounds", "outputs/wvcp_all_mcts_greedy_no_bounds/random"),
        # ("MCTS+R", "outputs/wvcp_all_mcts_greedy/random"),
        # ("MCTS+C no bounds", "outputs/wvcp_all_mcts_greedy_no_bounds/constrained"),
        # ("MCTS+C", "outputs/wvcp_all_mcts_greedy/constrained"),
        # ("MCTS+D no bounds", "outputs/wvcp_all_mcts_greedy_no_bounds/deterministic"),
        # ("MCTS+D", "outputs/wvcp_all_mcts_greedy/deterministic"),
        #
        # WVCP greedy vs MCTS+greedy
        # ("R", "outputs/wvcp_all_greedy/random"),
        # ("MCTS+R", "outputs/wvcp_all_mcts_greedy/random"),
        # ("C", "outputs/wvcp_all_greedy/constrained"),
        # ("MCTS+GR", "outputs/wvcp_all_mcts_greedy/constrained"),
        # ("D", "outputs/wvcp_all_greedy/deterministic"),
        # ("MCTS+G", "outputs/wvcp_all_mcts_greedy/deterministic"),
        # ("DSatur", "outputs/wvcp_all_greedy/dsatur"),
        # ("MCTS+DSatur", "outputs/wvcp_all_mcts_greedy/dsatur"),
        # ("RLF", "outputs/wvcp_all_greedy/rlf"),
        # ("MCTS+RLF", "outputs/wvcp_all_mcts_greedy/rlf"),
        # ("AFISA", "outputs/wvcp_all_ls/afisa_original"),
        # ("TabuWeight", "outputs/wvcp_all_ls/tabu_weight"),
        # ("RedLS", "outputs/wvcp_all_ls/redls"),
        # ("ILSTS", "outputs/wvcp_all_ls/ilsts"),
        #
        # GCP greedy vs MCTS+greedy vs NRPA
        # ("R", "outputs/gcp_all_greedy/random"),
        # ("MCTS+R", "outputs/gcp_all_mcts_greedy/random"),
        # ("C", "outputs/gcp_all_greedy/constrained"),
        # ("MCTS+C", "outputs/gcp_all_mcts_greedy/constrained"),
        # ("D", "outputs/gcp_all_greedy/deterministic"),
        # ("MCTS+D", "outputs/gcp_all_mcts_greedy/deterministic"),
        # ("DSatur", "outputs/gcp_all_greedy/DSatur"),
        # ("MCTS+DSatur", "outputs/gcp_all_mcts_greedy/dsatur"),
        # ("RLF", "outputs/gcp_all_greedy/RLF"),
        # ("MCTS+RLF", "outputs/gcp_all_mcts_greedy/rlf"),
        # ("NRPA", "outputs/gcp_all_NRPA"),
        # ("TabuCol", "outputs/gcp_all_tabu_col/tabu_col_optimized"),
        #
        # WVCP MCTS+LS variation of the time spent in LS
        # ("MCTS+AFISA_0.01", "outputs/wvcp_all_mcts_ls_0.01/afisa_original_0.01"),
        # ("MCTS+AFISA_0.02", "outputs/wvcp_all_mcts_ls_0.02/afisa_original_0.02"),
        # ("MCTS+AFISA_0.04", "outputs/wvcp_all_mcts_ls_0.04/afisa_original_0.04"),
        # ("MCTS+AFISA_0.08", "outputs/wvcp_all_mcts_ls_0.08/afisa_original_0.08"),
        # ("MCTS+AFISA_0.1", "outputs/wvcp_all_mcts_ls_0.1/afisa_original_0.1"),
        # ("MCTS+AFISA_0.2", "outputs/wvcp_all_mcts_ls_0.2/afisa_original_0.2"),
        # ("MCTS+TW_0.01", "outputs/wvcp_all_mcts_ls_0.01/tabu_weight_0.01"),
        # ("MCTS+TW_0.02", "outputs/wvcp_all_mcts_ls_0.02/tabu_weight_0.02"),
        # ("MCTS+TW_0.04", "outputs/wvcp_all_mcts_ls_0.04/tabu_weight_0.04"),
        # ("MCTS+TW_0.08", "outputs/wvcp_all_mcts_ls_0.08/tabu_weight_0.08"),
        # ("MCTS+TW_0.1", "outputs/wvcp_all_mcts_ls_0.1/tabu_weight_0.1"),
        # ("MCTS+TW_0.2", "outputs/wvcp_all_mcts_ls_0.2/tabu_weight_0.2"),
        # ("MCTS+RedLS_0.01", "outputs/wvcp_all_mcts_ls_0.01/redls_0.01"),
        # ("MCTS+RedLS_0.02", "outputs/wvcp_all_mcts_ls_0.02/redls_0.02"),
        # ("MCTS+RedLS_0.04", "outputs/wvcp_all_mcts_ls_0.04/redls_0.04"),
        # ("MCTS+RedLS_0.08", "outputs/wvcp_all_mcts_ls_0.08/redls_0.08"),
        # ("MCTS+RedLS_0.1", "outputs/wvcp_all_mcts_ls_0.1/redls_0.1"),
        # ("MCTS+RedLS_0.2", "outputs/wvcp_all_mcts_ls_0.2/redls_0.2"),
        # ("MCTS+ILSTS_0.01", "outputs/wvcp_all_mcts_ls_0.01/ilsts_0.01"),
        # ("MCTS+ILSTS_0.02", "outputs/wvcp_all_mcts_ls_0.02/ilsts_0.02"),
        # ("MCTS+ILSTS_0.04", "outputs/wvcp_all_mcts_ls_0.04/ilsts_0.04"),
        # ("MCTS+ILSTS_0.08", "outputs/wvcp_all_mcts_ls_0.08/ilsts_0.08"),
        # ("MCTS+ILSTS_0.1", "outputs/wvcp_all_mcts_ls_0.1/ilsts_0.1"),
        # ("MCTS+ILSTS_0.2", "outputs/wvcp_all_mcts_ls_0.2/ilsts_0.2"),
        #
        # WVCP LS vs MCTS+LS
        ("MCTS+GR", "outputs/wvcp_all_mcts_greedy/constrained"),
        ("MCTS+DSatur", "outputs/wvcp_all_mcts_greedy/dsatur"),
        ("AFISA", "outputs/wvcp_all_ls/afisa_original"),
        ("MCTS+AFISA", "outputs/wvcp_all_mcts_ls_0.02/afisa_original_0.02"),
        ("TabuWeight", "outputs/wvcp_all_ls/tabu_weight"),
        ("MCTS+TW", "outputs/wvcp_all_mcts_ls_0.02/tabu_weight_0.02"),
        ("RedLS", "outputs/wvcp_all_ls/redls"),
        ("MCTS+RedLS", "outputs/wvcp_all_mcts_ls_0.2/redls_0.2"),
        ("ILSTS", "outputs/wvcp_all_ls/ilsts"),
        ("MCTS+ILSTS", "outputs/wvcp_all_mcts_ls_0.02/ilsts_0.02"),
        #
        # WVCP MCTS+LS vs MCTS+HH (add the lines above for full display)
        ("Random", "outputs/wvcp_all_mcts_hh/random"),
        ("Deleter", "outputs/wvcp_all_mcts_hh/deleter"),
        ("Roulette", "outputs/wvcp_all_mcts_hh/roulette_wheel"),
        ("UCB", "outputs/wvcp_all_mcts_hh/ucb"),
        ("Pursuit", "outputs/wvcp_all_mcts_hh/pursuit"),
        ("NN", "outputs/wvcp_all_mcts_hh/neural_net"),
    ]

    problem = "gcp"
    problem = "wvcp"

    # Choose the set of instances
    instances_set = ("instance_list_gcp", "all")
    instances_set = ("../gcp_hard", "hard")
    instances_set = ("../gcp_medium", "medium")
    instances_set = ("../gcp_easy", "easy")

    instances_set = ("../DIMACS_hard", "dimacs_hard")
    instances_set = ("../DIMACS_easy", "dimacs_easy")
    instances_set = ("../rxx", "rxx")
    instances_set = ("../pxx", "pxx")
    instances_set = ("instance_list_wvcp", "all")

    # output_file = f"xlsx_files/wvcp_{instances_set[1]}_mcts_greedy_bounds.xlsx"
    # output_file = f"xlsx_files/wvcp_{instances_set[1]}_mcts_greedy.xlsx"
    # output_file = f"xlsx_files/gcp_{instances_set[1]}_mcts_greedy_NRPA.xlsx"
    # output_file = f"xlsx_files/wvcp_{instances_set[1]}_mcts_time_ls.xlsx"
    output_file = f"xlsx_files/wvcp_{instances_set[1]}_mcts_ls.xlsx"
    output_file = f"xlsx_files/wvcp_{instances_set[1]}_mcts_hh.xlsx"


    with open(f"instances/{instances_set[0]}.txt", "r", encoding="utf8") as file:
        instances = [i[:-1] for i in file.readlines()]

    table = Table(methods=methods, instances=instances, problem=problem)
    table.to_xlsx(output_file)
    print(output_file)


class Method:
    def __init__(self, name: str, repertory: str, instance_name: str) -> None:
        self.name: str = name
        self.repertory: str = repertory
        self.scores: list[float] = []
        self.times: list[float] = []
        self.optimal: bool = False
        self.time_optimal: list[float] = []
        self.mean_score: float = 0
        self.best_score: int = 0
        self.mean_best_time: float = 0
        self.mean_optimal_time: float = 0
        self.nb_best: int = 0
        self.nb_runs: int = 0
        self.nb_turns = []

        iter_max = {
            "C2000.5": 23,
            "C2000.9": 23,
            "DSJC1000.1": 44,
            "DSJC1000.5": 44,
            "DSJC1000.9": 44,
            "DSJC500.1": 86,
            "DSJC500.5": 86,
            "DSJC500.9": 86,
            "DSJC250.1": 164,
            "DSJC250.5": 164,
            "DSJC250.9": 164,
            "DSJC125.5gb": 301,
            "DSJC125.5g": 301,
            "DSJC125.9gb": 20,
            "DSJC125.9g": 20,
            "flat1000_50_0": 44,
            "flat1000_60_0": 44,
            "flat1000_76_0": 44,
            "latin_square_10": 49,
            "le450_15a": 106,
            "le450_15b": 101,
            "le450_15c": 95,
            "le450_15d": 95,
            "le450_25a": 129,
            "le450_25b": 20,
            "le450_25c": 101,
            "le450_25d": 101,
            "queen10_10": 361,
            "queen10_10gb": 361,
            "queen10_10g": 20,
            "queen11_11": 361,
            "queen11_11gb": 361,
            "queen11_11g": 361,
            "queen12_12": 301,
            "queen12_12gb": 301,
            "queen12_12g": 301,
            "queen13_13": 258,
            "queen14_14": 226,
            "queen15_15": 181,
            "queen16_16": 164,
            "wap01a": 21,
            "wap02a": 21,
            "wap03a": 10,
            "wap04a": 10,
            "wap05a": 55,
            "wap06a": 52,
            "wap07a": 26,
            "wap08a": 26,
        }

        # load data
        # files = sorted(
        #     glob(f"{repertory}/{instance_name}_[0-9]*_[0-9]*.csv"),
        #     key=lambda f: int(re.sub(r"\D", "", f)),
        # )
        # if files:
        #     for file_name in files:
        #         data = pd.read_csv(file_name, comment="#")
        #         time_: int = int(data.time.iloc[-1])
        #         nb_uncolored_: int = int(data.nb_uncolored.iloc[-1])
        #         penalty_: int = int(data.penalty.iloc[-1])
        #         score_: int = int(data.nb_colors.iloc[-1])
        #         if penalty_ != 0 or nb_uncolored_ != 0:
        #             continue
        #         if (
        #             "nb total node" in data.columns
        #             and "nb current node" in data.columns
        #         ):
        #             nb_total_node: int = int(data["nb total node"].iloc[-1])
        #             nb_current_node: int = int(data["nb current node"].iloc[-1])
        #             if nb_total_node > 1 and nb_current_node <= 1:
        #                 self.optimal = True
        #         self.scores.append(score_)
        #         self.times.append(time_)
        # else:
        files = sorted(
            glob(f"{repertory}/{instance_name}_[0-9]*.csv"),
            key=lambda f: int(re.sub(r"\D", "", f)),
        )
        for file_name in files:
            data = pd.read_csv(file_name, comment="#")
            time_: int = int(data.time.iloc[-1])
            # nb_uncolored_: int = int(data.nb_uncolored.iloc[-1])
            # penalty_: int = int(data.penalty.iloc[-1])
            # score_: int = int(data.nb_colors.iloc[-1])
            # if penalty_ != 0 or nb_uncolored_ != 0:
            #     continue

            if data.penalty.iloc[-1] != 0:
                print("error", file_name)
            if time_ < 0:
                time_ = 0
            if self.name in {
                "HEAD+Random",
                "HEAD+Deleter",
                "HEAD+Roulette",
                "HEAD+UCB",
                "HEAD+Pursuit",
                "HEAD+NN",
            }:
                # data = data.loc[data.time <= 3600]
                data = data.loc[data.turn <= iter_max[instance_name]]

            score_: int = (
                int(data.score.iloc[-1])
                if "score" in data
                else int(data.nb_colors.iloc[-1])
            )
            if "turn" in data:
                nb_turns_: int = int(data.turn.iloc[-1])
            else:
                nb_turns_ = 1
            if "nb total node" in data.columns and "nb current node" in data.columns:
                nb_total_node: int = int(data["nb total node"].iloc[-1])
                nb_current_node: int = int(data["nb current node"].iloc[-1])
                if nb_total_node > 1 and nb_current_node <= 1:
                    self.optimal = True
                    lines = []
                    with open(file_name, "r", encoding="utf8") as file:
                        for line in file.readlines():
                            if line.startswith("#"):
                                lines.append(line[1::])
                    # convert date (with format "year-month-day hour:minutes:seconds") to timestamp
                    date_start = datetime.strptime(
                        lines[1].split(",")[0], "%Y-%m-%d %H:%M:%S"
                    )
                    # try:
                    date_end = datetime.strptime(lines[2][:-1], "%Y-%m-%d %H:%M:%S")
                    self.time_optimal.append((date_end - date_start).total_seconds())
                    # except ValueError:
                    #     print(file_name, lines[2])
            self.scores.append(score_)
            self.times.append(time_)
            self.nb_turns.append(nb_turns_)

        if not files:
            files = sorted(
                glob(f"{repertory}/{instance_name}_[0-9]*.txt"),
                key=lambda f: int(re.sub(r"\D", "", f)),
            )
            for file_name in files:
                with open(file_name, "r", encoding="utf8") as file:
                    try:
                        score_, time_ = file.readlines()[1].strip().split(",")
                    except ValueError:
                        print(file_name, file.readlines())
                        score_ = "OF"
                if score_ == "OF":
                    continue
                self.scores.append(int(score_))
                self.times.append(float(time_))
        # if no available data
        if not files or not self.scores or not self.times:
            self.scores = [float("inf")]
            self.times = [float("inf")]
        # get mean, min,...
        self.mean_score = round(statistics.mean(self.scores), 1)
        self.best_score = min(self.scores)
        self.mean_best_time = round(
            statistics.mean(
                time
                for time, score in zip(self.times, self.scores)
                if score == self.best_score
            ),
            1,
        )
        self.mean_optimal_time = round(
            statistics.mean(self.time_optimal if self.time_optimal else [float("inf")]),
            1,
        )
        self.nb_best = sum(1 for score in self.scores if score == self.best_score)
        self.nb_runs = len(self.scores)


class Gap:
    def __init__(self, m1: Method, m2: Method) -> None:
        inf = float("inf")
        self.p_value: float = inf
        self.mean_score_difference: float = inf
        self.mean_best_time_difference: float = inf
        self.better_method_score: str = ""
        self.better_method_time: str = ""
        if m1.scores != [inf] and m2.scores != [inf]:
            if m1.mean_score == m1.best_score and m2.mean_score == m2.mean_score:
                p_value = 0
            else:
                # _, p_value = stats.ttest_ind(m1.scores, m2.scores)
                try:
                    _, p_value = stats.wilcoxon(m1.scores, m2.scores)
                except ValueError as e:
                    if sorted(m1.scores) == sorted(m2.scores):
                        pass
                    else:
                        print("_" * 50)
                        print(
                            m1.name,
                            sorted(m1.scores),
                            m2.name,
                            sorted(m2.scores),
                            e,
                            sep="\n",
                        )
                        print("-" * 50)
                    p_value = 0

            self.p_value = round(p_value, 3)

            self.mean_score_difference = round(
                m1.mean_score - m2.mean_score,
                1,
            )

            self.mean_best_time_difference = round(
                m1.mean_best_time - m2.mean_best_time, 1
            )

            if self.mean_score_difference < 0 and self.p_value <= P_VALUE_1:
                self.better_method_score = m1.name
            elif self.mean_score_difference > 0 and self.p_value <= P_VALUE_1:
                self.better_method_score = m2.name

            if self.mean_score_difference == 0 and self.mean_best_time_difference < -1:
                self.better_method_time = m1.name
            elif self.mean_score_difference == 0 and self.mean_best_time_difference > 1:
                self.better_method_time = m2.name


class Instance:
    """Store the results of all given method on one instance"""

    def __init__(
        self,
        name: str,
        methods: list[tuple[str, str]],
        problem: str,
        gaps: list[tuple[str, str]],
    ) -> None:
        print(name)
        self.name: str = name
        self.nb_vertices: int
        self.nb_edges: int
        self.best_known_score: int
        self.optimal: bool
        self.methods: dict[str, Method] = {
            m_name: Method(m_name, repertory, name) for m_name, repertory in methods
        }
        self.methods_names: list[str] = [m_name for m_name, _ in methods]
        self.gaps: dict[tuple[str, str], Gap] = {}
        self.best_found: int = min(m.best_score for m in self.methods.values())
        self.best_mean: float = min(m.mean_score for m in self.methods.values())

        # get informations on the instance
        self.nb_vertices, self.nb_edges = get_nb_vertices_edges(name)
        self.best_known_score, self.optimal = get_best_known_score(name, problem)

        # compute gap between methods
        for m1, m2 in gaps:
            self.gaps[(m1, m2)] = Gap(self.methods[m1], self.methods[m2])


class Table:
    """Representation of the data table"""

    def __init__(
        self,
        methods: list[tuple[str, str]],
        instances: list[str],
        problem: str,
    ) -> None:
        self.methods_names: list[str] = [m_name for m_name, _ in methods]

        self.gaps: list[tuple[str, str]] = []
        # pylint: disable=C0200
        for i in range(len(self.methods_names)):
            for j in range(i + 1, len(self.methods_names)):
                self.gaps.append((self.methods_names[i], self.methods_names[j]))

        self.instances: list[Instance] = [
            Instance(instance, methods, problem, self.gaps) for instance in instances
        ]
        self.nb_best_score = {
            m: sum(
                1
                for instance in self.instances
                if instance.best_known_score == instance.methods[m].best_score
            )
            for m in self.methods_names
        }
        self.nb_optim = {
            m: sum(1 for instance in self.instances if instance.methods[m].optimal)
            for m in self.methods_names
        }
        self.nb_best_found = {
            m: sum(
                1
                for instance in self.instances
                if instance.best_found == instance.methods[m].best_score
            )
            for m in self.methods_names
        }
        self.nb_best_mean = {
            m: sum(
                1
                for instance in self.instances
                if instance.best_mean == instance.methods[m].mean_score
            )
            for m in self.methods_names
        }

        self.nb_gaps_m1_best_score: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_score == m1)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m2_best_score: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_score == m2)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m1_best_time: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_time == m1)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }
        self.nb_gaps_m2_best_time: dict[tuple[str, str], int] = {
            (m1, m2): sum(
                int(instance.gaps[(m1, m2)].better_method_time == m2)
                for instance in self.instances
            )
            for (m1, m2) in self.gaps
        }

        self.ranks: dict[str, list[list[str]]] = dict()
        self.nb_ranks: dict[str, list[int]] = {
            m: [0] * len(self.methods_names) for m in self.methods_names
        }
        for instance in self.instances:
            methods_ = list(self.methods_names)
            keys = {
                m: (
                    instance.methods[m].mean_score,
                    instance.methods[m].best_score,
                    instance.methods[m].mean_best_time,
                )
                for m in methods_
            }
            methods_.sort(key=lambda m: keys[m])
            ranks: list[list[str]] = [[] for _ in methods_]
            last_rank = 0
            for i, m in enumerate(methods_):
                if not ranks[last_rank]:
                    ranks[last_rank].append(m)
                elif keys[ranks[last_rank][0]] == keys[m]:
                    ranks[last_rank].append(m)
                else:
                    last_rank = i
                    ranks[last_rank].append(m)
                self.nb_ranks[m][last_rank] += 1
            self.ranks[instance.name] = ranks

    def __repr__(self) -> str:
        return "\n".join([str(instance) for instance in self.instances])

    def table_results(self, workbook: Workbook):
        sheet = workbook.active
        sheet.title = "results"
        # first row
        # first columns are the instances informations then the methods names
        instance_info = ["instance", "|V|", "|E|", "BKS", "optim"]
        columns_info = ["best", "avg", "time", "optim time", "#"]
        line: list[int | str | float] = list(instance_info)
        line += [m for m in self.methods_names for _ in columns_info]
        sheet.append(line)
        # merge first row for methods names
        for i in range(len(self.methods_names)):
            sheet.merge_cells(
                start_row=1,
                end_row=1,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # second row
        # instance informations then columns info
        line = list(instance_info)
        for _ in self.methods_names:
            line += list(columns_info)
        sheet.append(line)

        # merge 2 firsts lines for instances informations
        for i in range(len(instance_info)):
            sheet.merge_cells(
                start_row=1, end_row=2, start_column=i + 1, end_column=i + 1
            )

        # body of the table
        # first columns are the instance info then the scores, times,... for each methods
        for instance in self.instances:
            line = [
                instance.name,
                instance.nb_vertices,
                instance.nb_edges,
                instance.best_known_score,
                instance.optimal,
            ]
            for m in self.methods_names:
                method = instance.methods[m]
                line += [
                    method.best_score,
                    method.mean_score,
                    method.mean_best_time,
                    method.mean_optimal_time
                    if method.mean_optimal_time != float("inf")
                    else "",
                    f"{method.nb_best}/{method.nb_runs}"
                    if method.scores != [float("inf")]
                    else "0/0",
                ]
            sheet.append(line)
            for col, m in enumerate(self.methods_names):
                column_best_score = len(instance_info) + 1 + len(columns_info) * col
                cell_best_score = sheet.cell(sheet.max_row, column_best_score)
                column_mean = column_best_score + 1
                cell_mean = sheet.cell(sheet.max_row, column_mean)
                if cell_best_score.value == float("inf"):
                    continue
                val_best_score = int(cell_best_score.value)
                if val_best_score == instance.best_known_score:
                    cell_best_score.font = Font(bold=True, color=COLOR_BEST)
                elif val_best_score < instance.best_known_score:
                    cell_best_score.font = Font(bold=True, color=COLOR_NEW_BEST)
                elif val_best_score == instance.best_found:
                    cell_best_score.font = Font(bold=True, color=COLOR_BEST_FOUND)
                if instance.methods[m].optimal:
                    cell_best_score.font = Font(bold=True, color=COLOR_OPTIMAL)
                val_best_mean = float(cell_mean.value)
                if instance.best_mean == val_best_mean:
                    cell_mean.font = Font(bold=True, color=COLOR_BEST_FOUND)

        # footer of the table with the number of best scores, ...

        # nb best scores
        line = ["#BKS"] * len(instance_info)
        for m in self.methods_names:
            line += [f"{self.nb_best_score[m]}/{len(self.instances)}"] * len(
                columns_info
            )
        sheet.append(line)
        # merge footer
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=1,
            end_column=len(instance_info),
        )
        for i, _ in enumerate(self.methods_names):
            sheet.merge_cells(
                start_row=sheet.max_row,
                end_row=sheet.max_row,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # nb best found
        line = ["#Best"] * len(instance_info)
        for m in self.methods_names:
            line += [f"{self.nb_best_found[m]}/{len(self.instances)}"] * len(
                columns_info
            )
        sheet.append(line)
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=1,
            end_column=len(instance_info),
        )
        for i, _ in enumerate(self.methods_names):
            sheet.merge_cells(
                start_row=sheet.max_row,
                end_row=sheet.max_row,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # nb best mean
        line = ["#Best Avg"] * len(instance_info)
        for m in self.methods_names:
            line += [f"{self.nb_best_mean[m]}/{len(self.instances)}"] * len(
                columns_info
            )
        sheet.append(line)
        # merge footer
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=1,
            end_column=len(instance_info),
        )

        for i, _ in enumerate(self.methods_names):
            sheet.merge_cells(
                start_row=sheet.max_row,
                end_row=sheet.max_row,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # nb optimal
        line = ["#Optimal"] * len(instance_info)
        for m in self.methods_names:
            line += [f"{self.nb_optim[m]}/{len(self.instances)}"] * len(columns_info)
        sheet.append(line)
        # merge footer
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=1,
            end_column=len(instance_info),
        )
        for i, _ in enumerate(self.methods_names):
            sheet.merge_cells(
                start_row=sheet.max_row,
                end_row=sheet.max_row,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # nb optimal
        line = ["#Same Score Reached"] * len(instance_info)
        for m in self.methods_names:
            nb_times_reach = 0
            nb_runs = 0
            for instance in self.instances:
                nb_times_reach += instance.methods[m].nb_best
                nb_runs += instance.methods[m].nb_runs
            line += [f"{nb_times_reach}/{nb_runs}"] * len(columns_info)
        sheet.append(line)
        # merge footer
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=1,
            end_column=len(instance_info),
        )
        for i, _ in enumerate(self.methods_names):
            sheet.merge_cells(
                start_row=sheet.max_row,
                end_row=sheet.max_row,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set optimal width
        column_widths: list[int] = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) + 1 > column_widths[i]:
                        column_widths[i] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for i, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = column_width

        # Freeze row and columns
        sheet.freeze_panes = sheet["F3"]

    def table_gaps(self, workbook: Workbook):
        sheet = workbook.create_sheet("gaps")

        instance_info = ["instance"]
        columns_info = ["gap score", "p_value", "gap time"]
        # first row
        line: list[int | str | float] = list(instance_info)
        line += [f"gap {m1} - {m2}" for m1, m2 in self.gaps for _ in columns_info]
        sheet.append(line)
        # merge first row for methods names
        for i in range(len(self.gaps)):
            sheet.merge_cells(
                start_row=1,
                end_row=1,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # second row
        # instance informations then columns info
        line = list(instance_info)
        for i in range(len(self.gaps)):
            line += list(columns_info)
        sheet.append(line)

        # merge 2 firsts lines for instances informations
        for i in range(len(instance_info)):
            sheet.merge_cells(
                start_row=1, end_row=2, start_column=i + 1, end_column=i + 1
            )

        # body of the table
        # first columns are the instance info then the gaps for each methods
        for instance in tqdm(self.instances):
            line = [instance.name]
            for m1, m2 in self.gaps:
                gap = instance.gaps[(m1, m2)]
                line += [
                    gap.mean_score_difference,
                    gap.p_value,
                    gap.mean_best_time_difference,
                ]
            sheet.append(line)

            # add color
            for col in range(len(self.gaps)):
                column_diff_score = len(instance_info) + 1 + len(columns_info) * col
                column_p_value = column_diff_score + 1
                column_diff_time = column_p_value + 1
                cell_diff_score = sheet.cell(sheet.max_row, column_diff_score)
                cell_p_value = sheet.cell(sheet.max_row, column_p_value)
                cell_diff_time = sheet.cell(sheet.max_row, column_diff_time)
                if cell_diff_score.value == float("inf"):
                    continue
                val_diff_score = float(cell_diff_score.value)
                val_p_value = float(cell_p_value.value)
                val_diff_time = float(cell_diff_time.value)

                color_score = "808080"  # gray
                color_time = "808080"  # gray
                if val_diff_score > 0:
                    # 1 better
                    if val_p_value <= P_VALUE_1:
                        color_score = COLOR_GAP1_1
                    elif val_p_value <= P_VALUE_2:
                        color_score = COLOR_GAP1_2
                    elif val_p_value <= P_VALUE_3:
                        color_score = COLOR_GAP1_3
                elif val_diff_score < 0:
                    # 2 better
                    if val_p_value <= P_VALUE_1:
                        color_score = COLOR_GAP2_1
                    elif val_p_value <= P_VALUE_2:
                        color_score = COLOR_GAP2_2
                    elif val_p_value <= P_VALUE_3:
                        color_score = COLOR_GAP2_3
                else:
                    if val_diff_time > 1:
                        color_time = COLOR_GAP1_2
                    elif val_diff_time < -1:
                        color_time = COLOR_GAP2_2
                cell_diff_score.font = Font(bold=True, color=color_score)
                cell_p_value.font = Font(bold=True, color=color_score)
                cell_diff_time.font = Font(bold=True, color=color_time)

        # footer
        line = ["better"] * len(instance_info)
        for m1, m2 in self.gaps:
            line += [
                f"{m1}",
                self.nb_gaps_m1_best_score[(m1, m2)],
                self.nb_gaps_m1_best_time[(m1, m2)],
            ]
        sheet.append(line)
        for col in range(len(self.gaps)):
            sheet.cell(
                sheet.max_row, len(instance_info) + 1 + len(columns_info) * col
            ).font = Font(bold=True, color=COLOR_GAP2_1)

        line = ["better"] * len(instance_info)
        for m1, m2 in self.gaps:
            line += [
                f"{m2}",
                self.nb_gaps_m2_best_score[(m1, m2)],
                self.nb_gaps_m2_best_time[(m1, m2)],
            ]
        sheet.append(line)
        for col in range(len(self.gaps)):
            sheet.cell(
                sheet.max_row, len(instance_info) + 1 + len(columns_info) * col
            ).font = Font(bold=True, color=COLOR_GAP1_1)

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set optimal width
        column_widths: list[int] = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) + 1 > column_widths[i]:
                        column_widths[i] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for i, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = column_width

        # Freeze row and columns
        sheet.freeze_panes = sheet["B3"]

    def table_comparaison_p_value(self, workbook: Workbook):
        sheet = workbook.create_sheet("score")
        # first row with the methods names
        line: list[int | str | float] = [""]
        for m1 in self.methods_names:
            line.append(m1)
        sheet.append(line)

        # add data, first method name then nb significative gaps
        for row, m1 in enumerate(self.methods_names):
            m1 = self.methods_names[row]
            line = []
            line.append(m1)
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    line.append("-")
                else:
                    if row < col:
                        line.append(self.nb_gaps_m1_best_score[(m1, m2)])
                    else:
                        line.append(self.nb_gaps_m2_best_score[(m2, m1)])
            sheet.append(line)

            # bold best scores
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    continue
                cell_m1 = sheet.cell(row + 2, col + 2)
                if (
                    row < col
                    and (
                        self.nb_gaps_m1_best_score[(m1, m2)]
                        > self.nb_gaps_m2_best_score[(m1, m2)]
                    )
                ) or (
                    row > col
                    and (
                        self.nb_gaps_m2_best_score[(m2, m1)]
                        > self.nb_gaps_m1_best_score[(m2, m1)]
                    )
                ):
                    cell_m1.font = Font(bold=True)

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Rotate the first row
        for col in range(len(self.methods_names)):
            met_name_cell = sheet.cell(1, col + 2)
            met_name_cell.alignment = Alignment(textRotation=90)

        # add info
        line = [""] * (len(self.methods_names) + 1)
        line.append(
            "number of times the method on the row is better than the method in the column"
        )
        sheet.append(line)

        # Set optimal width
        column_widths: list[int] = []
        for row_ in sheet:
            for col, cell in enumerate(row_):
                if len(column_widths) > col:
                    if len(str(cell.value)) + 1 > column_widths[col]:
                        column_widths[col] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for row, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(row)].width = column_width

    def table_comparaison_time(self, workbook: Workbook):
        sheet = workbook.create_sheet("time")
        # first row with the methods names
        line: list[str | int] = [""]
        for m1 in self.methods_names:
            line.append(m1)
        sheet.append(line)

        # add data, first method name then nb significative gaps
        for row, m1 in enumerate(self.methods_names):
            m1 = self.methods_names[row]
            line = []
            line.append(m1)
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    line.append("-")
                else:
                    if row < col:
                        line.append(self.nb_gaps_m1_best_time[(m1, m2)])
                    else:
                        line.append(self.nb_gaps_m2_best_time[(m2, m1)])
            sheet.append(line)

            # bold best scores
            for col, m2 in enumerate(self.methods_names):
                if m1 == m2:
                    continue
                cell_m1 = sheet.cell(row + 2, col + 2)
                if (
                    row < col
                    and (
                        self.nb_gaps_m1_best_time[(m1, m2)]
                        > self.nb_gaps_m2_best_time[(m1, m2)]
                    )
                ) or (
                    row > col
                    and (
                        self.nb_gaps_m2_best_time[(m2, m1)]
                        > self.nb_gaps_m1_best_time[(m2, m1)]
                    )
                ):
                    cell_m1.font = Font(bold=True)

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Rotate the first row
        for col in range(len(self.methods_names)):
            met_name_cell = sheet.cell(1, col + 2)
            met_name_cell.alignment = Alignment(textRotation=90)

        # add info
        line = [""] * (len(self.methods_names) + 1)
        line.append(
            "number of times the method on the row is better than the method in \n"
            "the column regarding the time if the score is the same (1s difference at least)"
        )
        sheet.append(line)

        # Set optimal width
        column_widths: list[int] = []
        row_: Iterable[Iterable[Any]]
        for row_ in sheet:
            for col, cell in enumerate(row_):
                if len(column_widths) > col:
                    if len(str(cell.value)) + 1 > column_widths[col]:
                        column_widths[col] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for row, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(row)].width = column_width

    def table_ranks(self, workbook: Workbook):
        sheet = workbook.create_sheet("ranks")
        # first row
        line = [""]
        for i in range(len(self.methods_names)):
            line.append(f"rank {i+1}")
        sheet.append(line)

        # data
        for instance in self.instances:
            line = [instance.name]
            for rank in self.ranks[instance.name]:
                txt = ""
                for m in rank:
                    txt += f"{m}\n"
                if rank:
                    m = rank[0]
                    txt += (
                        f"ms={instance.methods[m].mean_score}"
                        f" bs={instance.methods[m].best_score}"
                        f" mt={instance.methods[m].mean_best_time}"
                    )
                line.append(txt)
            sheet.append(line)

        # Set alignment
        for col in sheet.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set optimal width
        column_widths: list[int] = []
        for row in sheet:
            for row, cell in enumerate(row):
                if len(column_widths) > row:
                    if len(str(cell.value)) + 1 > column_widths[row]:
                        column_widths[row] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for row, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(row)].width = column_width

    def table_summary_ranks(self, workbook: Workbook):
        sheet = workbook.create_sheet("summary_ranks")
        # first row
        line: list[int | str | float] = [""]
        for i in range(len(self.methods_names)):
            line.append(f"rank {i+1}")
        sheet.append(line)

        methods = list(self.methods_names)
        scores = {m: 0 for m in methods}
        for m in methods:
            for i, nb in enumerate(self.nb_ranks[m]):
                scores[m] += nb * (len(self.nb_ranks[m]) - i)
        methods.sort(key=lambda m: scores[m], reverse=True)
        # data
        for m in methods:
            line = [m]
            for nb_rank in self.nb_ranks[m]:
                line.append(nb_rank)
            sheet.append(line)

        # Set alignment
        for col in sheet.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set optimal width
        column_widths: list[int] = []
        for row in sheet:
            for row, cell in enumerate(row):
                if len(column_widths) > row:
                    if len(str(cell.value)) + 1 > column_widths[row]:
                        column_widths[row] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for row, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(row)].width = column_width

    def table_results_article(self, workbook: Workbook):
        sheet = workbook.create_sheet("results article")
        # first row
        # first columns are the instances informations then the methods names
        instance_info = [
            "instance",
            # "|V|",
            # "|E|",
            "BKS",
        ]
        columns_info = ["best", "mean", "time"]
        line: list[int | str | float] = list(instance_info)
        line += [m for m in self.methods_names for _ in columns_info]
        sheet.append(line)
        # merge first row for methods names
        for i in range(len(self.methods_names)):
            sheet.merge_cells(
                start_row=1,
                end_row=1,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # second row
        # instance informations then columns info
        line = list(instance_info)
        for _ in self.methods_names:
            line += list(columns_info)
        sheet.append(line)

        # merge 2 firsts lines for instances informations
        for i in range(len(instance_info)):
            sheet.merge_cells(
                start_row=1, end_row=2, start_column=i + 1, end_column=i + 1
            )
        base_font = sheet["A1"].font
        # body of the table
        # first columns are the instance info then the scores, times,... for each methods
        for instance in self.instances:
            line = [
                instance.name,
                # instance.nb_vertices,
                # instance.nb_edges,
                f"{instance.best_known_score}*"
                if instance.optimal
                else instance.best_known_score,
            ]
            for m in self.methods_names:
                method = instance.methods[m]
                # the line is the best score, the mean score and the mean time of best scores
                # the best score is in bold and red if it is the best known score
                # the best score is in bold and green if it is a new best found score
                # the best score is in bold and blue if it is proved optimal
                # the mean score is not print if it is the same as the best score
                # cell_score = None
                # cell_mean_score = ""
                # if method.mean_score != method.best_score:
                #     if method.mean_score == instance.best_mean:
                #         cell_mean_score = TextBlock(
                #             font=InlineFont(b=True, color=COLOR_BEST_FOUND),
                #             text=f"~{method.mean_score}",
                #         )
                #     else:
                #         cell_mean_score = TextBlock(
                #             font=InlineFont(),
                #             text=f"~{method.mean_score}",
                #         )
                # if method.optimal:
                #     cell_score = CellRichText(
                #         [
                #             TextBlock(
                #                 font=InlineFont(b=True, color=COLOR_OPTIMAL),
                #                 text=f"{method.best_score}*",
                #             ),
                #             cell_mean_score,
                #         ]
                #     )
                # elif method.best_score == instance.best_known_score:
                #     cell_score = CellRichText(
                #         [
                #             TextBlock(
                #                 font=InlineFont(b=True, color=COLOR_BEST),
                #                 text=f"{method.best_score}",
                #             ),
                #             cell_mean_score,
                #         ]
                #     )
                # elif method.best_score < instance.best_known_score:
                #     cell_score = CellRichText(
                #         [
                #             TextBlock(
                #                 font=InlineFont(b=True, color=COLOR_NEW_BEST),
                #                 text=f"{method.best_score}",
                #             ),
                #             cell_mean_score,
                #         ]
                #     )
                # elif method.best_score == instance.best_found:
                #     cell_score = CellRichText(
                #         [
                #             TextBlock(
                #                 font=InlineFont(b=True, color=COLOR_BEST_FOUND),
                #                 text=f"{method.best_score}",
                #             ),
                #             cell_mean_score,
                #         ]
                #     )
                # else:
                #     cell_score = CellRichText(
                #         [
                #             TextBlock(
                #                 font=InlineFont(),
                #                 text=f"{method.best_score}",
                #             ),
                #             cell_mean_score,
                #         ]
                #     )

                # line += [
                #     cell_score,
                #     int(method.mean_best_time),
                # ]
                line += [
                    f"{method.best_score}*" if method.optimal else method.best_score,
                    method.mean_score
                    if method.mean_score != method.best_score
                    and method.mean_score != float("inf")
                    else "",
                    int(method.mean_best_time)
                    if method.mean_best_time != float("inf")
                    else "",
                ]
            sheet.append(line)
            for col, m in enumerate(self.methods_names):
                method = instance.methods[m]
                column_best_score = len(instance_info) + 1 + len(columns_info) * col
                cell_best_score = sheet.cell(sheet.max_row, column_best_score)
                column_mean = column_best_score + 1
                cell_mean = sheet.cell(sheet.max_row, column_mean)
                if method.best_score == float("inf"):
                    continue
                if method.best_score == instance.best_known_score:
                    cell_best_score.font = Font(bold=True, color=COLOR_BEST)
                elif method.best_score < instance.best_known_score:
                    cell_best_score.font = Font(bold=True, color=COLOR_NEW_BEST)
                elif method.best_score == instance.best_found:
                    cell_best_score.font = Font(bold=True, color=COLOR_BEST_FOUND)
                if method.optimal:
                    cell_best_score.font = Font(bold=True, color=COLOR_OPTIMAL)
                if (
                    method.mean_score == float("inf")
                    or method.mean_score == method.best_score
                ):
                    continue
                if instance.best_mean == method.mean_score:
                    cell_mean.font = Font(bold=True, color=COLOR_BEST_FOUND)

        # footer of the table with the number of best scores, ...

        # nb best scores
        line = ["#BKS"] * len(instance_info)
        for m in self.methods_names:
            line += [f"{self.nb_best_score[m]}/{len(self.instances)}"] * len(
                columns_info
            )
        sheet.append(line)
        # merge footer
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=1,
            end_column=len(instance_info),
        )
        for i, _ in enumerate(self.methods_names):
            sheet.merge_cells(
                start_row=sheet.max_row,
                end_row=sheet.max_row,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # nb best found
        line = ["#Best"] * len(instance_info)
        for m in self.methods_names:
            line += [f"{self.nb_best_found[m]}/{len(self.instances)}"] * len(
                columns_info
            )
        sheet.append(line)
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=1,
            end_column=len(instance_info),
        )
        for i, _ in enumerate(self.methods_names):
            sheet.merge_cells(
                start_row=sheet.max_row,
                end_row=sheet.max_row,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # nb best mean
        line = ["#Best Avg"] * len(instance_info)
        for m in self.methods_names:
            line += [f"{self.nb_best_mean[m]}/{len(self.instances)}"] * len(
                columns_info
            )
        sheet.append(line)
        # merge footer
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=1,
            end_column=len(instance_info),
        )

        for i, _ in enumerate(self.methods_names):
            sheet.merge_cells(
                start_row=sheet.max_row,
                end_row=sheet.max_row,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # nb optimal
        line = ["#Optimal"] * len(instance_info)
        for m in self.methods_names:
            line += [f"{self.nb_optim[m]}/{len(self.instances)}"] * len(columns_info)
        sheet.append(line)
        # merge footer
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=1,
            end_column=len(instance_info),
        )
        for i, _ in enumerate(self.methods_names):
            sheet.merge_cells(
                start_row=sheet.max_row,
                end_row=sheet.max_row,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # nb optimal
        line = ["#Same Score Reached"] * len(instance_info)
        for m in self.methods_names:
            nb_times_reach = 0
            nb_runs = 0
            for instance in self.instances:
                nb_times_reach += instance.methods[m].nb_best
                nb_runs += instance.methods[m].nb_runs
            line += [f"{nb_times_reach}/{nb_runs}"] * len(columns_info)
        sheet.append(line)
        # merge footer
        sheet.merge_cells(
            start_row=sheet.max_row,
            end_row=sheet.max_row,
            start_column=1,
            end_column=len(instance_info),
        )
        for i, _ in enumerate(self.methods_names):
            sheet.merge_cells(
                start_row=sheet.max_row,
                end_row=sheet.max_row,
                start_column=len(instance_info) + 1 + len(columns_info) * i,
                end_column=len(instance_info) + len(columns_info) * (i + 1),
            )

        # Set alignment
        for col_ in sheet.columns:
            for cell in col_:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set optimal width
        column_widths: list[int] = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) + 1 > column_widths[i]:
                        column_widths[i] = len(str(cell.value)) + 1
                else:
                    column_widths += [0]
        for i, column_width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = column_width

        # Freeze row and columns
        sheet.freeze_panes = sheet["C3"]

    def to_xlsx(self, file_name: str):
        """Convert the table to xlsx file"""
        workbook = Workbook()

        # first sheet with all results of each methods
        print("generation results")
        self.table_results(workbook)

        # second sheet with the gaps between each methods
        print("generation gaps")
        self.table_gaps(workbook)

        # summary of the gaps on the score with the p value
        print("generation comparison p value")
        self.table_comparaison_p_value(workbook)

        # summary of the gaps on the times
        print("generation comparison time")
        self.table_comparaison_time(workbook)

        # rank of the methods for each instances
        print("generation ranks")
        self.table_ranks(workbook)

        # summary of the ranks for each methods
        print("generation summary ranks")
        self.table_summary_ranks(workbook)

        # results for article
        self.table_results_article(workbook)
        workbook.save(file_name)


def get_nb_vertices_edges(instance: str) -> tuple[int, int]:
    """return nb vertices and nb edges"""
    with open("instances/instance_info.txt", "r", encoding="utf8") as file:
        for line in file.readlines():
            instance_, nb_vertices, nb_edges = line[:-1].split(",")
            if instance_ == instance:
                return int(nb_vertices), int(nb_edges)
    raise Exception(f"instance {instance} not found in instances/instance_info.txt")


def get_best_known_score(instance: str, problem: str) -> tuple[int, bool]:
    """return best know score in the literature and if score optimal"""
    file: str = f"instances/best_scores_{problem}.txt"
    with open(file, "r", encoding="utf8") as f:
        for line in f.readlines():
            instance_, score, optimal = line[:-1].split(" ")
            if instance_ == instance:
                return int(score), optimal == "*"
    raise Exception(f"instance {instance} not found in {file}")


if __name__ == "__main__":
    main()
